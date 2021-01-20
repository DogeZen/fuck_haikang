# encoding=utf8
from selenium import webdriver
from selenium.common.exceptions import ElementNotVisibleException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.select import Select
import time
import win32gui
import openpyxl
import pyautogui
import ctypes, sys


def is_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False


# 前台开启浏览器模式
def openChrome():
    # 加启动配置
    option = webdriver.ChromeOptions()
    # option.add_argument('--headless')
    option.add_argument('disable-infobars')

    # 打开chrome浏览器
    driver = webdriver.Chrome(chrome_options=option)
    driver.maximize_window()
    return driver





# 授权操作
def operationAuth(driver):
    url = "http://192.168.1.64/"
    driver.get(url)

    try:
        elem_pw = driver.find_elements_by_xpath("//input[@type='password']")
        for elem in elem_pw:
            elem.send_keys("Buchou123")
        # class="aui_state_highlight" class="aui_close"  placeholder="用户名"
        driver.find_element_by_xpath("//button[@class='aui_state_highlight']").click()
        time.sleep(1)
        driver.find_element_by_xpath("//div[@class='aui_titleBar']//a").click()
    except ElementNotVisibleException:
        print("已注册")
        driver.find_element_by_xpath("//input[@type='password']").clear()
    try:
        driver.find_element_by_xpath("//input[@type='password']").send_keys("Buchou123")
        driver.find_element_by_xpath("//input[@placeholder='用户名']").send_keys("admin")
        driver.find_element_by_xpath("//button").click()
    except:
        pass
    print("登陆成功")
    time.sleep(2)
    driver.find_element_by_xpath("//a[@ng-bind='oLan.config']").click()
    record_project(driver)
    advanced_setting(driver)
    osd(driver)
    region_cover(driver)
    smart_detection(driver)
    storage_manage(driver)
    video_setting(driver)
    record_device_num(driver)


def advanced_setting(driver):
    driver.find_element_by_xpath("//span[@class='menu-icon network-icon']").click()
    driver.find_element_by_xpath("//span[@ng-bind='oMenuLan.advancedSettings']").click()
    time.sleep(2)
    print("选中高级设置")
    driver.find_element_by_xpath('//li[@module="advancedPlatform"]//a').click()
    time.sleep(2)
    driver.find_element_by_xpath("//select[@ng-model='oParams.szAccessType']").click()
    driver.find_element_by_xpath("//option[@value='ezviz']").click()
    time.sleep(1)
    # if driver.find_element_by_xpath("//input[@ng-model='oParams.bEnableEzviz']").is_selected() != True:
    driver.find_element_by_xpath("//input[@ng-model='oParams.bEnableEzviz']").click()
    time.sleep(1)
    driver.find_element_by_xpath("//input[@ng-model='oParams.bEnableEzviz']").click()
    try:
        time.sleep(1)
        driver.find_elements_by_xpath("//input")[0].clear()
        time.sleep(1)
        driver.find_elements_by_xpath("//input")[1].clear()
        time.sleep(1)
        driver.find_elements_by_xpath("//input")[0].send_keys('12341234q')
        driver.find_elements_by_xpath("//input")[1].send_keys('12341234q')
        driver.find_element_by_xpath("//button[@class='aui_state_highlight']").click()
    except:
        pass
    time.sleep(1)
    driver.find_element_by_xpath('//input[@input-valid="oParamsEzvizValid.oVerifyCode"]').clear()
    driver.find_element_by_xpath('//input[@input-valid="oParamsEzvizValid.oVerifyCode"]').send_keys('12341234q')
    driver.find_element_by_xpath("//button[@class='btn btn-primary btn-save']").click()
    print("萤石云配置成功")
def osd(driver):
    driver.find_element_by_xpath('//span[@ng-bind="oMenuLan.image"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//a[@ng-bind="oImageLan.OSDSettings"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//input[@ng-model="oOsdParams.cameraName"]').clear()
    book = openpyxl.load_workbook('养户摄像头对应表.xlsx')
    sheet = book.worksheets[0]

    for column in sheet.columns:
        # print(column[0].value)
        if column[0].value == '序列号':
            for i in range(len(column)):
                if column[i].value is None:
                    devices_num = i
                    # print("已录入设备数目为" + str(devices_num - 1))
                    break
    for column in sheet.columns:
        if column[0].value == 'osd':
            name = column[devices_num].value
            # print(name)
    if name is not None:

        driver.find_element_by_xpath('//input[@ng-model="oOsdParams.cameraName"]').send_keys(name)
        driver.find_element_by_xpath('//button[@ng-click="save()"]').click()
        print("osd修改完成")
    else:
        print("养户名为空，未修改osd！")


# 遮挡区域配置
def region_cover(driver):
    driver.find_element_by_xpath('//span[@class="menu-icon event-icon"]').click()
    time.sleep(2)
    driver.find_element_by_xpath('//a[@ng-bind="oLan.videoTamper"]').click()
    time.sleep(2)
    # print("选中图像")
    if driver.find_element_by_xpath('//input[@id="enableVideoTamper"]').is_selected() != True:
        driver.find_element_by_xpath('//input[@id="enableVideoTamper"]').click()
    driver.find_element_by_xpath('//button[@ng-click="clear()"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//button[@class="aui_state_highlight"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//button[@ng-click="draw()"]').click()
    time.sleep(1)
    hwnd_title = {}

    def get_all_hwnd(hwnd, mouse):
        if (win32gui.IsWindow(hwnd) and
                win32gui.IsWindowEnabled(hwnd) and
                win32gui.IsWindowVisible(hwnd)):
            hwnd_title.update({hwnd: win32gui.GetWindowText(hwnd)})

    win32gui.EnumWindows(get_all_hwnd, 0)

    for h, t in hwnd_title.items():
        if t:
            # print(h, t)
            if t == 'WindowControl':
                left, top, right, bottom = win32gui.GetWindowRect(h)
                # print(left, top, right, bottom)

                pyautogui.moveTo(left + 1, top + 1)
                time.sleep(1)
                pyautogui.dragTo(x=right - 1, y=bottom - 1, duration=0.1)
                time.sleep(1)
    slider = driver.find_element_by_xpath('//div[@class="slider"]')
    # print(slider.size['width'])
    action = ActionChains(driver)
    action.move_to_element(slider)
    # action.move_by_offset(slider.size['width']/2,0)
    action.click()
    action.perform()
    driver.find_element_by_xpath('//button[@class="btn btn-primary btn-save"]').click()
    print("绘制遮挡完成")
    driver.find_element_by_xpath('//div[@class ="step last"]').click()
    time.sleep(1)
    if driver.find_element_by_xpath('//input[@id="enableVideoTamper"]').is_selected() != True:
        driver.find_element_by_xpath('//input[@id="enableVideoTamper"]').click()
    # if driver.find_element_by_xpath('//input[@ng-click="alarmLinkAll()"]').is_selected() != True:
    #     driver.find_element_by_xpath('//input[@ng-click="alarmLinkAll()"]').click()
    if not driver.find_element_by_xpath('//input[@ng-click="normalLinkAll()"]').is_selected():
        driver.find_element_by_xpath('//input[@ng-click="normalLinkAll()"]').click()
    driver.find_element_by_xpath('//button[@ng-click="save()"]').click()
    print("联动方式设置完成")

def smart_detection(driver):
    #Smart事件

    driver.find_element_by_xpath('//span[@ng-bind="oMenuLan.smartEvent"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//a[@ng-bind="oLan.intrusionDetect"]').click()
    time.sleep(1)
    if not driver.find_element_by_css_selector('.checkbox.ng-pristine.ng-valid').is_selected():
        driver.find_element_by_css_selector('.checkbox.ng-pristine.ng-valid').click()
        time.sleep(1)
        driver.find_element_by_xpath('//button[@class="btn btn-primary btn-save"]').click()

        time.sleep(1)


def record_project(driver):
    driver.find_element_by_xpath('//span[@class ="menu-icon storage-icon"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//span[@ng-bind="oMenuLan.scheduleSettings"]').click()
    time.sleep(1)
    if driver.find_element_by_xpath('//input[@class="checkbox ng-pristine ng-valid"]').is_selected() != True:
        driver.find_element_by_xpath('//input[@class="checkbox ng-pristine ng-valid"]').click()
    driver.find_element_by_xpath('//span[@id="recordDiv_deleteAll_txt"]').click()

    sliders = driver.find_elements_by_xpath('//div[@class="timeplan_daytimeplan"]')
    for slider in sliders:
        # print(slider.size)
        action = ActionChains(driver)
        action.move_to_element(slider)
        action.move_by_offset(-slider.size['width'] / 2, 0)
        action.click_and_hold()
        action.move_by_offset(slider.size['width'], 0)
        action.release()
        # action.drag_and_drop_by_offset(slider,slider.size['width']/2,0)
        action.perform()
    driver.find_element_by_xpath('//button[@ng-click="save()"]').click()
    print("录像计划配置完成")


def video_setting(driver):
    time.sleep(2)
    driver.find_element_by_xpath('//span[@class="menu-icon video-icon"]').click()
    time.sleep(2)
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.resolution"]')).select_by_index(0)
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.bitrateType"]')).select_by_index(1)
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.videoQuality"]')).select_by_index(2)
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.frameRate"]')).select_by_value("11")
    driver.find_element_by_xpath('//input[@ng-model="oVideoParams.maxBitrate"]').clear()
    driver.find_element_by_xpath('//input[@ng-model="oVideoParams.maxBitrate"]').send_keys("2048")
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.videoEncoding"]')).select_by_index(1)
    driver.find_element_by_xpath('//button[@class="btn btn-primary btn-save"]').click()


def storage_manage(driver):
    # 存储配置
    driver.find_element_by_xpath('//span[@class ="menu-icon storage-icon"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//span[@ng-bind="oMenuLan.storageManagement"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//input[@ng-model="oParams.iVideoQuotaRatio"]').clear()
    driver.find_element_by_xpath('//input[@ng-model="oParams.iVideoQuotaRatio"]').send_keys("90")
    driver.find_element_by_xpath('//span[@ng-bind="oLan.save"]').click()
    driver.find_element_by_xpath('//input[@class="checkbox"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//input[@ng-value="oLan.format"]').click()
    time.sleep(1)
    try:
        driver.find_element_by_xpath('//button[@class="aui_state_highlight"]').click()
        print("检测到内存卡已插入，开始格式化")
    except:
        print("未插内存卡,插上内存卡重装一次")
        print("未插内存卡,插上内存卡重装一次")
        print("未插内存卡,插上内存卡重装一次")
        print("未插内存卡,插上内存卡重装一次")
        print("未插内存卡,插上内存卡重装一次")
        print("未插内存卡,插上内存卡重装一次")
        print("未插内存卡,插上内存卡重装一次")
        print("未插内存卡,插上内存卡重装一次")
        print("插上内存卡后重装")
        input()
        input()
        input()
        input()

    print("等待设备格式化完成，三分钟内请勿拔电源")
    # while(True):
    #     try:
    #         driver.find_element_by_xpath('//img[@src="../ui/images/artDialog/loading.gif"]')
    #     except:
    #         break
    driver.refresh()
    time.sleep(2)
    print("存储配置完成")


def record_device_num(driver):
    driver.find_element_by_xpath('//span[@class="menu-icon system-icon"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//span[@ng-bind="oMenuLan.systemSettings"]').click()
    time.sleep(1)
    device_num = driver.find_element_by_xpath('//input[@ng-model = "oSettingBasicInfo.szSerialNo"]').get_attribute(
        'value')[-9:]

    print("序列号为" + device_num)
    record_num(device_num)


def record_num(device_num):
    try:
        book = openpyxl.load_workbook('养户摄像头对应表.xlsx')
        sheet = book.worksheets[0]
        devices_column_index = 1
        for column in sheet.columns:

            # print(column[0].value)
            if (column[0].value == '序列号'):
                column_num = devices_column_index
                for i in range(len(column)):
                    if (column[i].value is None):
                        devices_num = i
                        print("该excel表中已录入设备数目为" + str(devices_num ))
                        break
            devices_column_index += 1
        sheet.cell(row=devices_num + 1, column=column_num, value=device_num)
        book.save('养户摄像头对应表.xlsx')
        book.close()
    except:
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("关闭后按任意键继续")
        input()
        record_num(device_num)


# 方法主入口
if __name__ == '__main__':

    if is_admin():
        print("已使用管理员权限打开")
        # 加启动配置
        driver = openChrome()
        print('''
                请确认以下事项:
                1.画遮挡区域时不要移动鼠标直到配置完成
                2.配置过程中不要打开excel，否则会导致无法写入
                ''')
        operationAuth(driver)
        while True:
            print('''
                ————————————————
                ————————————————
                按任意键进入下一台配置''')
            input()
            operationAuth(driver)
    else:
        print("未使用管理员权限打开，自动获取管理员权限")
        # Re-run the program with admin rights
        ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, __file__, None, 1)
